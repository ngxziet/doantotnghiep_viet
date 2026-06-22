
# -*- coding: utf-8 -*-
"""Sinh file Excel ket qua thuc nghiem (muc 5.2) voi du lieu mau + cong thuc song.
CANH BAO: so lieu trong file la MAU, phai thay bang do thuc truoc khi nop.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HDR = PatternFill("solid", fgColor="305496")
HDRF = Font(bold=True, color="FFFFFF")
NOTE = PatternFill("solid", fgColor="FFF2CC")
RES = PatternFill("solid", fgColor="E2EFDA")   # o cong thuc (ket qua)
CEN = Alignment(horizontal="center", vertical="center", wrap_text=True)
B = Side(style="thin", color="BFBFBF")
BORDER = Border(left=B, right=B, top=B, bottom=B)


def style_header(ws, row, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR; cell.font = HDRF; cell.alignment = CEN; cell.border = BORDER


def grid(ws, r0, r1, ncol):
    for r in range(r0, r1 + 1):
        for c in range(1, ncol + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = CEN; cell.border = BORDER


def banner(ws, ncol, text):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    c = ws.cell(row=1, column=1, value=text)
    c.fill = NOTE; c.font = Font(bold=True, color="9C5700"); c.alignment = CEN
    ws.row_dimensions[1].height = 28


def widths(ws, ws_widths):
    for i, w in enumerate(ws_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


wb = Workbook()
NOTE_TXT = "SO LIEU MAU - THAY BANG DO THUC TE TRUOC KHI NOP. O xanh la = cong thuc tu tinh, dung sua."

# ---------- 5.1 Quang duong ----------
ws = wb.active; ws.title = "5.1 Quang duong"
banner(ws, 9, NOTE_TXT)
hdr = ["Quang lenh (m)", "Lan1", "Lan2", "Lan3", "Lan4", "Lan5",
       "Do thuc TB (m)", "Sai so TB (cm)", "Sai so (%)"]
ws.append(hdr); style_header(ws, 2, 9)
data = [
    (0.5, [0.48, 0.49, 0.47, 0.48, 0.48]),
    (1.0, [0.97, 0.96, 0.98, 0.97, 0.97]),
    (2.0, [1.95, 1.94, 1.96, 1.95, 1.95]),
]
for cmd, trials in data:
    r = ws.max_row + 1
    ws.cell(row=r, column=1, value=cmd)
    for i, v in enumerate(trials):
        ws.cell(row=r, column=2 + i, value=v)
    ws.cell(row=r, column=7, value=f"=AVERAGE(B{r}:F{r})").fill = RES
    ws.cell(row=r, column=8, value=f"=ABS(A{r}-G{r})*100").fill = RES
    ws.cell(row=r, column=9, value=f"=ABS(A{r}-G{r})/A{r}*100").fill = RES
grid(ws, 3, ws.max_row, 9)
widths(ws, [13, 8, 8, 8, 8, 8, 13, 13, 11])

# ---------- 5.2 Xoay goc ----------
ws = wb.create_sheet("5.2 Xoay goc")
banner(ws, 8, NOTE_TXT)
ws.append(["Goc lenh (do)", "Lan1", "Lan2", "Lan3", "Lan4", "Lan5",
           "Goc do TB (do)", "Sai so TB (do)"]); style_header(ws, 2, 8)
ws.append([]); ws.cell(row=3, column=10, value="Do lech chuan (do)").font = Font(bold=True)
data = [
    (45, [44.5, 43.8, 44.0, 45.0, 43.6]),
    (90, [88.5, 89.2, 87.6, 88.9, 88.8]),
    (180, [177.5, 176.0, 178.2, 175.5, 177.0]),
]
# header them cot do lech chuan
ws.cell(row=2, column=9, value="Do lech chuan (do)"); style_header(ws, 2, 9)
ws.delete_rows(3, 1)  # bo dong rong vua them
for cmd, trials in data:
    r = ws.max_row + 1
    ws.cell(row=r, column=1, value=cmd)
    for i, v in enumerate(trials):
        ws.cell(row=r, column=2 + i, value=v)
    ws.cell(row=r, column=7, value=f"=AVERAGE(B{r}:F{r})").fill = RES
    ws.cell(row=r, column=8, value=f"=ABS(A{r}-G{r})").fill = RES
    ws.cell(row=r, column=9, value=f"=STDEV.S(B{r}:F{r})").fill = RES
grid(ws, 3, ws.max_row, 9)
widths(ws, [13, 8, 8, 8, 8, 8, 14, 14, 16])

# ---------- 5.3 Giu huong 3m ----------
ws = wb.create_sheet("5.3 Giu huong")
banner(ws, 3, NOTE_TXT)
ws.append(["Lan do", "Do lech ngang tai 3m (cm)", "Goc lech tuong duong (do)"])
style_header(ws, 2, 3)
devs = [6, 9, 5]
for i, d in enumerate(devs, start=1):
    r = ws.max_row + 1
    ws.cell(row=r, column=1, value=i)
    ws.cell(row=r, column=2, value=d)
    ws.cell(row=r, column=3, value=f"=DEGREES(ATAN(B{r}/300))").fill = RES
r = ws.max_row + 1
ws.cell(row=r, column=1, value="Trung binh").font = Font(bold=True)
ws.cell(row=r, column=2, value=f"=AVERAGE(B3:B{r-1})").fill = RES
ws.cell(row=r, column=3, value=f"=AVERAGE(C3:C{r-1})").fill = RES
grid(ws, 3, ws.max_row, 3)
widths(ws, [12, 26, 26])

# ---------- 5.4 Node ----------
ws = wb.create_sheet("5.4 Node")
banner(ws, 5, NOTE_TXT)
ws.append(["Lo trinh", "So node", "Tong quang (m)",
           "Sai so vi tri cuoi (cm)", "Sai so goc cuoi (do)"]); style_header(ws, 2, 5)
rows = [
    ("Duong thang 4 node", 4, 2.0, 7.0, 2.5),
    ("Hinh vuong canh 0.5m", 4, 2.0, 12.0, 4.8),
    ("Chu L", 3, 1.5, 9.0, 3.6),
]
for row in rows:
    ws.append(list(row))
grid(ws, 3, ws.max_row, 5)
widths(ws, [22, 9, 14, 22, 20])

# ---------- 5.5 HC-SR04 ----------
ws = wb.create_sheet("5.5 HC-SR04")
banner(ws, 5, NOTE_TXT)
ws.append(["Khoang cach thuc (cm)", "Do giua (cm)", "Do trai +90 (cm)",
           "Do phai -90 (cm)", "Sai so TB (cm)"]); style_header(ws, 2, 5)
rows = [
    (10, 10.2, 10.4, 9.8),
    (20, 20.3, 19.6, 20.5),
    (30, 29.7, 30.6, 29.4),
    (50, 50.8, 49.1, 51.2),
]
for real, g, t, p in rows:
    r = ws.max_row + 1
    ws.cell(row=r, column=1, value=real)
    ws.cell(row=r, column=2, value=g)
    ws.cell(row=r, column=3, value=t)
    ws.cell(row=r, column=4, value=p)
    ws.cell(row=r, column=5,
            value=f"=(ABS(B{r}-A{r})+ABS(C{r}-A{r})+ABS(D{r}-A{r}))/3").fill = RES
grid(ws, 3, ws.max_row, 5)
widths(ws, [20, 13, 16, 16, 14])

# ---------- 5.6 Ne vat can ----------
ws = wb.create_sheet("5.6 Ne vat can")
banner(ws, 4, NOTE_TXT)
ws.append(["Luot thu", "Khoang cach dung (cm)", "Ne thanh cong (Co/Khong)", "Ghi chu"])
style_header(ws, 2, 4)
rows = [
    (1, 28.5, "Co", ""),
    (2, 29.2, "Co", ""),
    (3, 27.8, "Co", ""),
    (4, 26.5, "Khong", "Vat mong (chan ghe), phat hien tre"),
    (5, 29.0, "Co", ""),
]
for row in rows:
    ws.append(list(row))
last = ws.max_row
r = last + 1
ws.cell(row=r, column=1, value="Ti le thanh cong").font = Font(bold=True)
ws.cell(row=r, column=2,
        value=f'=COUNTIF(C3:C{last},"Co")&"/"&COUNTA(C3:C{last})&" ("&'
              f'ROUND(COUNTIF(C3:C{last},"Co")/COUNTA(C3:C{last})*100,0)&"%)"').fill = RES
r2 = r + 1
ws.cell(row=r2, column=1, value="Kc dung TB (cm)").font = Font(bold=True)
ws.cell(row=r2, column=2, value=f"=AVERAGE(B3:B{last})").fill = RES
grid(ws, 3, ws.max_row, 4)
ws.cell(row=last, column=4).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
widths(ws, [12, 22, 24, 34])

# ---------- 5.7 Drift ----------
ws = wb.create_sheet("5.7 Drift yaw")
banner(ws, 2, NOTE_TXT)
ws.append(["Thoi gian dung yen (s)", "Do troi yaw (do)"]); style_header(ws, 2, 2)
for t, d in [(30, 0.3), (60, 0.6), (300, 2.1)]:
    ws.append([t, d])
grid(ws, 3, ws.max_row, 2)
widths(ws, [22, 18])

out = "docs/ket-qua-thuc-nghiem-5.2.xlsx"
wb.save(out)
print("Saved:", out, "| sheets:", wb.sheetnames)
